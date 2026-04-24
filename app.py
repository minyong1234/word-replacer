from flask import Flask, request, send_file, render_template, jsonify
from replacer import replace_in_docx
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
import os, json, tempfile, zipfile, copy, io, re, calendar
from datetime import date, timedelta
from concurrent.futures import ThreadPoolExecutor

app = Flask(__name__)

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "template.xlsx")

# ── 템플릿 고정 상수 ──────────────────────────────────────────────────────────
GANTT_ROW_START  = 6    # 간트 시작 행
GANTT_ROW_END    = 22   # 간트 끝 행 (취약점 자체점검 포함)
GANTT_COL_START  = 6    # 간트 시작 열 (F열)
HEADER_ROW_YEAR  = 3    # 연도 헤더 행
HEADER_ROW_MONTH = 4    # 월 헤더 행
HEADER_ROW_WEEK  = 5    # 주 헤더 행
VULN_ROW         = 22   # 취약점 자체점검 행
LAST_MS_ROW      = 21   # 마지막 마일스톤 행

# 템플릿 기준: F열(col6) = 4월 1주
TMPL_YEAR        = 2026
TMPL_START_MONTH = 4
TMPL_GANTT_START_COL = 6  # 템플릿에서 간트 첫 컬럼

# ── 헬퍼 함수 ────────────────────────────────────────────────────────────────
def is_colored(f):
    if not f or f.fill_type != 'solid': return False
    ft = f.fgColor.type
    if ft == 'theme' and f.fgColor.theme != 0: return True
    if ft == 'rgb':
        try: return f.fgColor.rgb not in ('00000000', 'FFFFFFFF', '00FFFFFF')
        except: pass
    return False

def get_week_count(year, month):
    """첫 번째 월요일이 있는 주=1주, 마지막 날이 속한 주=마지막 주 (최대 5주)"""
    first_day = date(year, month, 1)
    last_day  = date(year, month, calendar.monthrange(year, month)[1])
    days_to_monday = (7 - first_day.weekday()) % 7
    first_monday   = first_day if days_to_monday == 0 else first_day + timedelta(days=days_to_monday)
    if first_monday > last_day:
        return 0
    last_monday = last_day - timedelta(days=last_day.weekday())
    return min((last_monday - first_monday).days // 7 + 1, 5)

def next_ym(year, month):
    return (year + 1, 1) if month == 12 else (year, month + 1)

def scan_gantt(ws_src):
    """템플릿에서 색칠된 셀을 자동 스캔 → {row: [col, ...]}"""
    gantt = {}
    for rn in range(GANTT_ROW_START, GANTT_ROW_END + 1):
        for col in range(GANTT_COL_START, ws_src.max_column + 1):
            c = ws_src.cell(rn, col)
            if isinstance(c, MergedCell): continue
            if is_colored(c.fill):
                gantt.setdefault(rn, []).append(col)
    return gantt

def build_col_map(start_year, start_month):
    """
    실제 시작월 기준으로 col=6부터 (year, month) 매핑 생성.
    최대 36달치 미리 계산.
    """
    col = GANTT_COL_START
    mapping = {}  # col → (year, month)
    cur_y, cur_m = start_year, start_month
    for _ in range(36):
        wc = get_week_count(cur_y, cur_m)
        for _ in range(wc):
            mapping[col] = (cur_y, cur_m)
            col += 1
        cur_y, cur_m = next_ym(cur_y, cur_m)
    return mapping

def build_header_layout(start_year, start_month, end_year, end_month):
    """
    col=6부터 시작월 1주 ~ end_month까지 헤더 레이아웃 생성.
    header는 항상 시작월 1주부터 표기.
    간트 색칠 시작 위치는 SHIFT(=시작주 0-indexed)로 별도 제어.
    """
    col, layout = GANTT_COL_START, []
    cur_y, cur_m = start_year, start_month
    while (cur_y, cur_m) <= (end_year, end_month):
        wc = get_week_count(cur_y, cur_m)
        for w in range(1, wc + 1):
            layout.append({
                'col': col, 'year': cur_y, 'month': cur_m, 'week': w,
                'is_month_start': w == 1, 'is_month_end': w == wc,
            })
            col += 1
        cur_y, cur_m = next_ym(cur_y, cur_m)
    return layout


def generate_wbs(client_name, start_date_str, include_vuln_self):
    sd           = date.fromisoformat(start_date_str)
    start_year   = sd.year
    start_month  = sd.month

    # SHIFT = (시작일의 주차 0-indexed) - 1
    # 템플릿 원본 col7이 "시작 다음주"이므로 -1 보정해서 "시작주 자체"로 맞춤
    # 예) 5월 18일(3주) → swo=2, SHIFT=1 → col7+1=col8=5월3주 색칠 시작 ✓
    # 예) 5월 6일(1주)  → swo=0, SHIFT=-1 → col7-1=col6=5월1주 색칠 시작 ✓
    first_day_of_month = sd.replace(day=1)
    days_to_monday = (7 - first_day_of_month.weekday()) % 7
    first_monday = (first_day_of_month if days_to_monday == 0
                    else first_day_of_month + timedelta(days=days_to_monday))
    swo = (sd - first_monday).days // 7 if sd >= first_monday else 0
    SHIFT = swo - 1

    med    = Side(border_style="medium")
    thin   = Side(border_style="thin")
    none_s = Side(border_style=None)

    wb     = load_workbook(TEMPLATE_PATH)
    ws     = wb.active
    wb_src = load_workbook(TEMPLATE_PATH)
    ws_src = wb_src.active

    # ── 고객사명 치환 ─────────────────────────────────────────────────────────
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell): continue
            if cell.value and isinstance(cell.value, str) and "고객사명" in cell.value:
                cell.value = cell.value.replace("고객사명", client_name)

    # ── 간트 색상 스캔 (템플릿 자동 인식) ────────────────────────────────────
    gantt = scan_gantt(ws_src)

    # ── 간트 초기화 (색상 + 테두리 전체 클리어) ─────────────────────────────
    for rn in range(GANTT_ROW_START, GANTT_ROW_END + 1):
        for col in range(GANTT_COL_START, ws.max_column + 1):
            c = ws.cell(rn, col)
            if isinstance(c, MergedCell): continue
            try:
                if is_colored(c.fill): c.fill = PatternFill(fill_type=None)
            except: pass
            c.border = Border()  # 테두리도 전부 클리어

    # ── 간트 shift 적용 ───────────────────────────────────────────────────────
    for rn, cols in gantt.items():
        if rn == VULN_ROW: continue  # 취약점은 별도 처리
        for oc in cols:
            new_col = oc + SHIFT
            if new_col < GANTT_COL_START: continue
            target = ws.cell(rn, new_col)
            if not isinstance(target, MergedCell):
                target.fill = copy.copy(ws_src.cell(rn, oc).fill)

    # ── 취약점 자체점검 행 (22행) shift ──────────────────────────────────────
    if include_vuln_self and VULN_ROW in gantt:
        for oc in gantt[VULN_ROW]:
            new_col = oc + SHIFT
            if new_col < GANTT_COL_START: continue
            target = ws.cell(VULN_ROW, new_col)
            if not isinstance(target, MergedCell):
                target.fill = copy.copy(ws_src.cell(VULN_ROW, oc).fill)

    # ── 마지막 마일스톤 컬럼 → 실제 (year, month) 파악 ──────────────────────
    # 템플릿 기준 col_map으로 orig_col의 월 파악 후 start_month offset 적용
    last_ms_orig_col = max(gantt.get(LAST_MS_ROW, [GANTT_COL_START]))
    # 템플릿 기준 col → (year, month) 매핑
    _tmpl_col = GANTT_COL_START
    _tmpl_col_map = {}
    _cur_y, _cur_m = TMPL_YEAR, TMPL_START_MONTH
    for _ in range(36):
        _wc = get_week_count(_cur_y, _cur_m)
        for _w in range(1, _wc + 1):
            _tmpl_col_map[_tmpl_col] = (_cur_y, _cur_m, _w)
            _tmpl_col += 1
        _cur_y, _cur_m = next_ym(_cur_y, _cur_m)
    # 템플릿에서 last_ms_orig_col의 (월, 주) 파악
    _tmpl_y, _tmpl_m, _tmpl_w = _tmpl_col_map.get(last_ms_orig_col, (TMPL_YEAR, 12, 1))
    # 실제 연도/월: 템플릿 기준 월에서 start_month 차이만큼 이동
    _month_offset = (_tmpl_m - TMPL_START_MONTH) + (_tmpl_y - TMPL_YEAR) * 12
    _actual_months = (start_month - 1) + _month_offset
    last_ms_actual_year  = start_year + _actual_months // 12
    last_ms_actual_month = _actual_months % 12 + 1
    end_year, end_month = next_ym(last_ms_actual_year, last_ms_actual_month)

    # ── 병합 완전 해제 ────────────────────────────────────────────────────────
    for mr in [str(m) for m in ws.merged_cells.ranges]:
        ws.unmerge_cells(mr)

    # ── 헤더 행 기존 값/서식 클리어 (3,4,5행 간트 영역) ─────────────────────
    for rn in (HEADER_ROW_YEAR, HEADER_ROW_MONTH, HEADER_ROW_WEEK):
        for col in range(GANTT_COL_START, ws.max_column + 2):
            c = ws.cell(rn, col)
            if isinstance(c, MergedCell): continue
            c.value  = None
            c.fill   = PatternFill(fill_type=None)
            c.border = Border()

    # ── 헤더 레이아웃 생성 ────────────────────────────────────────────────────
    header_layout = build_header_layout(start_year, start_month, end_year, end_month)

    # 서식 참조 (원본 헤더 셀)
    ref3 = ws_src.cell(HEADER_ROW_YEAR,  GANTT_COL_START)
    ref4 = ws_src.cell(HEADER_ROW_MONTH, GANTT_COL_START)
    ref5 = ws_src.cell(HEADER_ROW_WEEK,  GANTT_COL_START)

    # 연도 구간 파악
    year_groups = {}
    for item in header_layout:
        y = item['year']
        if y not in year_groups:
            year_groups[y] = {'start': item['col'], 'end': item['col']}
        else:
            year_groups[y]['end'] = item['col']

    # 행별 top/bottom 테두리 참조 (원본 간트 영역 중간 컬럼)
    ref_border_col = GANTT_COL_START + 5
    row_borders = {}
    for rn in range(GANTT_ROW_START, GANTT_ROW_END + 1):
        b = ws_src.cell(rn, ref_border_col).border
        row_borders[rn] = {'top': b.top, 'bottom': b.bottom}

    # ── 헤더 작성 ─────────────────────────────────────────────────────────────
    for item in header_layout:
        col           = item['col']
        is_year_start = (col == year_groups[item['year']]['start'])
        is_year_end   = (col == year_groups[item['year']]['end'])

        # 5행: 주
        c5 = ws.cell(HEADER_ROW_WEEK, col)
        c5.value     = f"{item['week']}주"
        c5.font      = copy.copy(ref5.font)
        c5.fill      = copy.copy(ref5.fill)
        c5.alignment = copy.copy(ref5.alignment)
        c5.border    = Border(left=thin, right=thin, top=med, bottom=med)

        # 4행: 월
        c4 = ws.cell(HEADER_ROW_MONTH, col)
        c4.value     = f"{item['month']}월" if item['is_month_start'] else None
        c4.font      = copy.copy(ref4.font)
        c4.fill      = copy.copy(ref4.fill)
        c4.alignment = copy.copy(ref4.alignment)
        c4.border    = Border(
            left   = med if item['is_month_start'] else none_s,
            right  = med if item['is_month_end']   else none_s,
            top=med, bottom=med,
        )

        # 3행: 연도
        c3 = ws.cell(HEADER_ROW_YEAR, col)
        c3.value     = f"{item['year']}년" if is_year_start else None
        c3.font      = copy.copy(ref3.font)
        c3.fill      = copy.copy(ref3.fill)
        c3.alignment = copy.copy(ref3.alignment)
        c3.border    = Border(
            left   = med if is_year_start else none_s,
            right  = med if is_year_end   else none_s,
            top=med, bottom=none_s,
        )

        # 간트 행: thin 테두리
        for rn in range(GANTT_ROW_START, GANTT_ROW_END + 1):
            rb = row_borders.get(rn, {'top': thin, 'bottom': thin})
            # 취약점 자체점검 포함 시 22행 하단은 thin (굵은 테두리 제거)
            bottom = thin if (include_vuln_self and rn == VULN_ROW) else rb['bottom']
            ws.cell(rn, col).border = Border(
                left=thin, right=thin,
                top=rb['top'], bottom=bottom,
            )

        ws.column_dimensions[get_column_letter(col)].width = 4.9

    # ── 3행 제목 셀 서식 복원 (B3:E4) ────────────────────────────────────────
    title_cell = ws.cell(HEADER_ROW_YEAR, 2)
    title_cell.value = "간편등급 CSAP 컨설팅 일정"
    src_title = ws_src.cell(HEADER_ROW_YEAR, 2)
    title_cell.font      = copy.copy(src_title.font)
    title_cell.fill      = copy.copy(src_title.fill)
    title_cell.alignment = copy.copy(src_title.alignment)
    title_cell.border    = copy.copy(src_title.border)

    # ── 병합 재설정 ───────────────────────────────────────────────────────────
    def sm(sc, sr, ec, er):
        if sc <= ec and sr <= er:
            try: ws.merge_cells(start_row=sr, start_column=sc, end_row=er, end_column=ec)
            except: pass

    # 제목 병합 (B3:E4)
    sm(2, HEADER_ROW_YEAR, 5, HEADER_ROW_MONTH)

    # 4행: 월별 병합
    month_groups = {}
    for item in header_layout:
        key = (item['year'], item['month'])
        if key not in month_groups:
            month_groups[key] = {'start': item['col'], 'end': item['col']}
        else:
            month_groups[key]['end'] = item['col']
    for key, cols in month_groups.items():
        if cols['start'] < cols['end']:
            sm(cols['start'], HEADER_ROW_MONTH, cols['end'], HEADER_ROW_MONTH)

    # 3행: 연도별 병합
    for y, cols in year_groups.items():
        if cols['start'] < cols['end']:
            sm(cols['start'], HEADER_ROW_YEAR, cols['end'], HEADER_ROW_YEAR)

    # 좌측 항목 영역 고정 병합 (원본 그대로)
    for (sc, sr, ec, er) in [
        (2, 7, 2, 9), (3, 7, 3, 9), (4, 8, 4, 9),
        (2, 10, 2, 12), (3, 10, 3, 12), (4, 11, 4, 12),
        (3, 14, 3, 17), (3, 19, 3, 20),
    ]:
        try: ws.merge_cells(start_row=sr, start_column=sc, end_row=er, end_column=ec)
        except: pass

    # ── 전체 글자 검은색 ──────────────────────────────────────────────────────
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell): continue
            if cell.font:
                f = cell.font
                cell.font = Font(name=f.name, size=f.size, bold=f.bold,
                                 italic=f.italic, underline=f.underline,
                                 strike=f.strike, color="FF000000")

    # ── 취약점 자체점검 미포함 시 행 삭제 ────────────────────────────────────
    if not include_vuln_self:
        ws.delete_rows(VULN_ROW, 1)
        for col in range(2, GANTT_COL_START):
            c = ws.cell(VULN_ROW - 1, col)
            b = c.border
            c.border = Border(left=b.left, right=b.right, top=b.top, bottom=med)

    safe_name = re.sub(r'[\\/:*?"<>|]', '_', client_name)
    filename  = f"{safe_name}_CSAP_간편등급_컨설팅_일정_v2_1.xlsx"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, filename


# ── 워드 치환 관련 ────────────────────────────────────────────────────────────
def process_file(file_data):
    file_bytes, filename, rules, tmpdir = file_data
    input_path  = os.path.join(tmpdir, f"in_{filename}")
    output_path = os.path.join(tmpdir, f"out_{filename}")
    with open(input_path, "wb") as f:
        f.write(file_bytes)
    replace_in_docx(input_path, output_path, rules)
    return output_path, filename


# ── 라우트 ───────────────────────────────────────────────────────────────────
@app.route("/")
def home():
    return render_template("home.html")

@app.route("/replacer")
def replacer():
    return render_template("index.html")

@app.route("/replace", methods=["POST"])
def replace():
    if "files" not in request.files:
        return jsonify({"error": "파일이 없습니다"}), 400
    files = request.files.getlist("files")
    rules_json = request.form.get("rules", "{}")
    rules = json.loads(rules_json)
    if not rules:
        return jsonify({"error": "치환 규칙을 입력해주세요"}), 400

    with tempfile.TemporaryDirectory() as tmpdir:
        if len(files) == 1:
            file = files[0]
            if not file.filename.endswith(".docx"):
                return jsonify({"error": ".docx 파일만 지원합니다"}), 400
            input_path  = os.path.join(tmpdir, file.filename)
            output_path = os.path.join(tmpdir, f"out_{file.filename}")
            file.save(input_path)
            replace_in_docx(input_path, output_path, rules)
            return send_file(output_path, as_attachment=True, download_name=file.filename)
        else:
            file_data_list = []
            for file in files:
                if not file.filename.endswith(".docx"):
                    continue
                file_data_list.append((file.read(), file.filename, rules, tmpdir))
            with ThreadPoolExecutor() as executor:
                results = list(executor.map(process_file, file_data_list))
            zip_path = os.path.join(tmpdir, "replaced_files.zip")
            with zipfile.ZipFile(zip_path, "w") as zipf:
                for output_path, filename in results:
                    zipf.write(output_path, filename)
            return send_file(zip_path, as_attachment=True, download_name="replaced_files.zip")

@app.route("/wbs")
def wbs():
    return render_template("wbs.html")

@app.route("/wbs/generate", methods=["POST"])
def wbs_generate():
    client_name       = request.form.get("client_name", "").strip()
    start_date        = request.form.get("start_date", "")
    include_vuln_self = request.form.get("include_vuln_self") == "true"
    if not client_name or not start_date:
        return jsonify({"error": "고객사명과 시작일을 입력해주세요."}), 400
    try:
        buf, filename = generate_wbs(client_name, start_date, include_vuln_self)
        return send_file(buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=filename)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True)
